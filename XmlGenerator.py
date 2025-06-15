import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import datetime
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import numbers
from PIL import Image, ImageTk

class TemplateGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("Document Template Generator")
        self.root.geometry("700x950")
        self.build_form()

    def build_form(self):
        tk.Label(self.root, text="Malcode:").grid(row=0, column=0, sticky='e')
        self.malcode_entry = tk.Entry(self.root, width=50)
        self.malcode_entry.grid(row=0, column=1)

        tk.Label(self.root, text="Team 1:").grid(row=1, column=0, sticky='e')
        self.team1_entry = tk.Entry(self.root, width=50)
        self.team1_entry.grid(row=1, column=1)

        tk.Label(self.root, text="Team 2:").grid(row=2, column=0, sticky='e')
        self.team2_entry = tk.Entry(self.root, width=50)
        self.team2_entry.grid(row=2, column=1)

        tk.Label(self.root, text="Team 3:").grid(row=3, column=0, sticky='e')
        self.team3_entry = tk.Entry(self.root, width=50)
        self.team3_entry.grid(row=3, column=1)

        tk.Label(self.root, text="Team 4:").grid(row=4, column=0, sticky='e')
        self.team4_entry = tk.Entry(self.root, width=50)
        self.team4_entry.grid(row=4, column=1)

        tk.Label(self.root, text="Delivery Manager:").grid(row=5, column=0, sticky='e')
        self.manager_entry = tk.Entry(self.root, width=50)
        self.manager_entry.grid(row=5, column=1)

        tk.Label(self.root, text="Time Scheduled (HH:MM):").grid(row=6, column=0, sticky='e')
        self.initial_time_entry = tk.Entry(self.root, width=50)
        self.initial_time_entry.grid(row=6, column=1)

        tk.Label(self.root, text="Describe Each Task:").grid(row=7, column=0, sticky='nw', pady=(10, 0))
        self.task_descriptions = []
        self.task_labels = [
            "Implementation",
            "Implementation",
            "Validation",
            "Implementation",
            "Validation",
            "Go-No-Go",
            "Back Out"
        ]

        for i, label in enumerate(self.task_labels):
            tk.Label(self.root, text=f"Task {i+1}: {label}").grid(row=8+i, column=0, sticky='e')
            text = tk.Text(self.root, width=50, height=4)  # 4 lines tall
            text.grid(row=8 + i, column=1, pady=2)
            self.task_descriptions.append(text)

        self.generate_btn = tk.Button(self.root, text="Generate Template", command=self.generate_template)
        self.generate_btn.grid(row=15, column=1, pady=10)

        # Logo
        try:
            image = Image.open("logo.jpg")  # Place logo.jpg in the same directory
            max_width = 350
            aspect_ratio = image.height / image.width
            new_height = int(max_width * aspect_ratio)
            resized = image.resize((max_width, new_height))
            photo = ImageTk.PhotoImage(resized)
            self.image_label = tk.Label(self.root, image=photo)
            self.image_label.image = photo
            self.image_label.grid(row=16, column=0, columnspan=4, pady=14, sticky="n")

        except Exception as e:
            print(f"Logo could not be loaded: {e}")

    def generate_template(self):
        try:
            initial_time = datetime.datetime.strptime(self.initial_time_entry.get(), "%H:%M")
        except ValueError:
            messagebox.showerror("Invalid Time", "Initial time must be in format HH:MM")
            return

        data = [
            ["Malcode", self.malcode_entry.get(), "CONTACT"],
            ["Team 1", self.team1_entry.get(), "", ""],
            ["Team 2", self.team2_entry.get(), "", ""],
            ["Team 3", self.team3_entry.get(), "", ""],
            ["Team 4", self.team4_entry.get(), "", ""],
            ["Delivery Manager", self.manager_entry.get(), "", ""],
            ["TIME SCHEDULED", self.initial_time_entry.get(), "", ""],
            [],
            ["Deployment"],
            ["Time Start", "Time Finish", "Task Descriptions", "Time Required (min)", "Team Name", "Task"]
        ]

        base_tasks = [
            ["Team 1", "Implementation", 30],
            ["Team 1", "Implementation", 30],
            ["Team 2", "Validation", 20],
            ["Team 1", "Implementation", 30],
            ["Team 2", "Validation", 20],
            ["Team 3", "Review", 25],
            ["Team 1", "Back Out", 15]
        ]

        start_row_index = len(data) + 1
        task_rows = []

        for i, (team, task, duration) in enumerate(base_tasks):
            current_row = start_row_index + i
            start_cell = f"A{current_row}"

            if i == 0:
                time_start_formula = "=B7"
            else:
                prev_finish_cell = f"B{current_row - 1}"
                time_start_formula = f"={prev_finish_cell}"

            time_finish_formula = f"={start_cell}+TIME(0,{duration},0)"
            description = self.task_descriptions[i].get()

            task_rows.append([
                time_start_formula, time_finish_formula, description, duration, team, task
            ])

        data.extend(task_rows)

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        df = pd.DataFrame(data)
        df.to_excel(file_path, index=False, header=False)

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Set column widths
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20

        # Set number format for time
        for row in range(start_row_index, start_row_index + len(base_tasks)):
            ws[f"A{row}"].number_format = "HH:MM"
            ws[f"B{row}"].number_format = "HH:MM"

        # Fill colors
        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        purple = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")
        orange = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")

        ws["A9"].fill = yellow  # "Deployment"
        for cell in ws[10]:  # Header row
            cell.fill = yellow
        for cell in ws[11]:  # First task row (optional highlight)
            cell.fill = green

        for i in range(1, 8):
            ws[f"A{i}"].fill = green
            ws[f"B{i}"].fill = purple
            ws[f"C{i}"].fill = orange

        wb.save(file_path)
        messagebox.showinfo("Success", f"Template saved to: {file_path}")

# -------------------- Run the GUI --------------------
if __name__ == '__main__':
    root = tk.Tk()
    app = TemplateGenerator(root)
    root.mainloop()
